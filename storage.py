from __future__ import annotations

import sqlite3
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# (excel_key, header). Особый ключ "yield_display" собирается из yield_pct
# или yield_formula на лету.
COLUMNS: list[tuple[str, str]] = [
    ("isin",                     "ISIN"),
    ("name",                     "Наименование"),
    ("yield_date",               "Дата, к которой рассчитана доходность"),
    ("yield_date_type",          "Тип даты"),
    ("yield_display",            "Доходность, %"),
    ("years_to_date",            "Лет до даты"),
    ("maturity_date",            "Дата погашения"),
    ("last_price_pct",           "Цена, % от номинала"),
    ("coupon_yield_pct",         "Купонная доходность, %"),
    ("current_yield_pct",        "Текущая купонная доходность, %"),
    ("coupon_quantity_per_year", "Купонов в год"),
    ("aci",                      "НКД"),
    ("aci_currency",             "Валюта НКД"),
    ("nominal",                  "Номинал"),
    ("nominal_currency",         "Номинал валюта"),
    ("risk_level",               "Уровень риска"),
    ("amortization",             "Амортизация"),
    ("perpetual",                "Бессрочные"),
    ("floater",                  "Плавающий купон"),
    ("figi",                     "FIGI"),
]


def init_db(path: Path) -> None:
    with sqlite3.connect(path) as con:
        con.execute("""
            CREATE TABLE IF NOT EXISTS snapshots (
                snapshot_ts              TEXT NOT NULL,
                isin                     TEXT NOT NULL,
                figi                     TEXT,
                name                     TEXT,
                maturity_date            TEXT,
                yield_date               TEXT,
                yield_date_type          TEXT,
                yield_pct                REAL,
                yield_formula            TEXT,
                years_to_date            REAL,
                last_price_pct           REAL,
                coupon_yield_pct         REAL,
                current_yield_pct        REAL,
                coupon_quantity_per_year INTEGER,
                aci                      REAL,
                aci_currency             TEXT,
                nominal                  REAL,
                nominal_currency         TEXT,
                risk_level               TEXT,
                amortization             INTEGER,
                perpetual                INTEGER,
                floater                  INTEGER,
                PRIMARY KEY (snapshot_ts, isin)
            )
        """)


def save_snapshot_sqlite(path: Path, rows: list[dict]) -> None:
    if not rows:
        return
    with sqlite3.connect(path) as con:
        con.executemany("""
            INSERT OR REPLACE INTO snapshots (
                snapshot_ts, isin, figi, name, maturity_date,
                yield_date, yield_date_type, yield_pct, yield_formula, years_to_date,
                last_price_pct, coupon_yield_pct, current_yield_pct,
                coupon_quantity_per_year,
                aci, aci_currency, nominal, nominal_currency,
                risk_level, amortization, perpetual, floater
            ) VALUES (
                :snapshot_ts, :isin, :figi, :name, :maturity_date,
                :yield_date, :yield_date_type, :yield_pct, :yield_formula, :years_to_date,
                :last_price_pct, :coupon_yield_pct, :current_yield_pct,
                :coupon_quantity_per_year,
                :aci, :aci_currency, :nominal, :nominal_currency,
                :risk_level, :amortization, :perpetual, :floater
            )
        """, rows)


def _xl_value(row: dict, key: str):
    if key == "yield_display":
        v = row.get("yield_pct")
        if v is not None:
            return v
        return row.get("yield_formula")
    v = row.get(key)
    if isinstance(v, bool):
        return "да" if v else "нет"
    return v


def write_excel(path: Path, rows: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Облигации"

    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, (_, title) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.alignment = header_align

    for r_idx, row in enumerate(rows, 2):
        for c_idx, (key, _) in enumerate(COLUMNS, 1):
            ws.cell(row=r_idx, column=c_idx, value=_xl_value(row, key))

    pct_cols = {
        "yield_display", "coupon_yield_pct", "current_yield_pct", "last_price_pct",
    }
    decimal_cols = {"aci", "nominal", "years_to_date"}
    for c_idx, (key, _) in enumerate(COLUMNS, 1):
        if key in pct_cols or key in decimal_cols:
            for r_idx in range(2, len(rows) + 2):
                cell = ws.cell(row=r_idx, column=c_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"

    ws.freeze_panes = "A2"
    for c_idx, (key, title) in enumerate(COLUMNS, 1):
        max_len = len(title)
        for row in rows:
            v = _xl_value(row, key)
            if v is not None:
                s = f"{v:.2f}" if isinstance(v, float) else str(v)
                if len(s) > max_len:
                    max_len = len(s)
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max_len + 2, 50)

    wb.save(path)
